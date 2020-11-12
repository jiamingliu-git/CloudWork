from openpyxl import load_workbook

#方法一：(需要用的时候，再调用这个类拿数据),优势：占用内存小，劣势：磁盘读写压力大

#拿到新建任务的数据
# class Getdata():
#     def __init__(self,fileName,sheetNum):
#         self.fileName=fileName
#         self.sheetNum=sheetNum
#
#     def cell(self,row,col):
#         wb = load_workbook(self.fileName)._sheets[self.sheetNum]
#         res=wb.cell(row,col).value
#         return res


class LoadFile():
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

#获取第一行：
    def get_header(self):
        wb=load_workbook(self.file_name)
        sheet=wb._sheets[self.sheet_name]
        header=[]
        for j in range(1,sheet.max_column+1):
            header.append(sheet.cell(1,j).value)
        return header                       #拿到第一行

#自动获取excel用例方法：
    def get_data_auto(self):
        wb = load_workbook(self.file_name)
        sheet = wb._sheets[self.sheet_name]
        header=self.get_header()
        testcase_data = []
        for i in range(2,sheet.max_row):
            sub_data={}
            for j in range(1,sheet.max_column+1):
                sub_data[header[j-1]]=sheet.cell(i,j).value
            testcase_data.append(sub_data)
        return testcase_data


#需要手动改：
    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet=wb._sheets[self.sheet_name]
        testcase_data = []                                         # 新建一个列表，把数据全部放进列表
        for i in range(2,sheet.max_row):                            #读第2到最大行数，读出来再切
            box = {}                                                 # 新建一个空字典，把数据放进字典
            box['CaseNum'] = sheet.cell(i,1).value
            box['Module'] = sheet.cell(i,2).value
            box['CaseName'] = sheet.cell(i,3).value
            box['Method'] = sheet.cell(i,4).value                    # 新建/定义空字典里addtest_url的值，值是我从excel里取的
            box['Headers'] = sheet.cell(i,5).value                   # 新建/定义空字典里addtest_data的值，值是我从excel里取的
            box['Url'] = sheet.cell(i,6).value                       # 新建/定义空字典里addtest_expect的值，值是我从excel里取的
            box['Data'] = sheet.cell(i,7).value                      # 新建一个列表，把数据全部放进列表
            box['Expect'] = sheet.cell(i,8).value
            box['Result'] = sheet.cell(i,9).value
            box['Form'] = sheet.cell(i,10).value
            box['Mention'] = sheet.cell(i,11).value

            testcase_data.append(box)
        return testcase_data

#用例执行结果写回excel：
    def write_back(self,i,value):
        wb=load_workbook(self.file_name)
        sheet=wb._sheets[self.sheet_name]
        sheet.cell(i,9).value=value                 #结果写在第九列的result里
        wb.save(self.file_name)

if __name__ == '__main__':
    print('*********')

#
# file='F:\pyfile\CloudWork\TestData\data_task.xlsx'
# res=DoExcel(file,0).get_data()
# print('手动',res)
# res1=DoExcel(file,0).get_data_auto()
# print('自动',res1)
# res=DoExcel(file,0).get_header()
# print(res)

# res=DoExcel('F:\pyfile\CloudWork\TestData\data_task.xlsx',0).get_data()[4]
# print(res)



##**************************************使用说明*****************************************************************

# file='datafile.xlsx'
# sheet1=load_workbook(file)._sheets[0]                       # 打开这个file文件的第一页，如果要打开第二页，sheets后面的0改为1
# max_row=sheet1.max_row
# max_column=sheet1.max_column
#
# print('excel文件datafile，第一页的名字：',sheet1)
# print('excel文件datafile，第一页的最大行数：',max_row)
# print('excel文件datafile，第一页的最大列数：：',max_column)
#
# a=sheet1.cell(1,1).value                                    # 定位这一页的第1行，第1列的内容
# b=sheet1.cell(2,1).value                                    # 定位这一页的第2行，第1列的内容
# print(a)
# print(b)




#******************************************使用说明*********************************************************
#

# #取登录的token值有三种方式：
# # setUp
# # global全局变量
# #反射：见下文
#
# from CloudWork_Project.MainCode.TestCase_Task import TestCase_Box_AddTask
#
# class GetToken:
#     login_token=None
# setattr(GetToken,'login_token',TestCase_Box_AddTask().setUp())
#
# print(GetToken.login_token)






# import xlrd
# import xlwt
# print(xlrd.open_workbook('datafile.xlsx').sheet_names())
#
# print(xlrd.open_workbook('datafile.xlsx').sheet_names()[0])
#
# print(xlrd.open_workbook('datafile.xlsx').sheet_by_name('第一页').row_values(1)[1])
#


